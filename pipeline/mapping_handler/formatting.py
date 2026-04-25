"""
mapping_handler/formatting.py — Excel formatting for mapped output
====================================================================
Professional styling applied to the effectiveness_mapped.xlsx workbook.

Edit this file to change header colours, fonts, column widths, etc.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def format_mapped_excel(path: str) -> None:
    """Apply professional styling to the mapped output workbook at *path*."""
    wb = load_workbook(path)
    ws = wb.active

    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font   = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h_fill   = PatternFill("solid", start_color="1F4E79")
    h_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d_font   = Font(name="Arial", size=9)
    d_align  = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", start_color="EBF3FB")

    # Header row
    ws.row_dimensions[1].height = 40
    for cell in ws[1]:
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = border

    # Data rows
    for ri in range(2, ws.max_row + 1):
        ws.row_dimensions[ri].height = 20
        fill = alt_fill if ri % 2 == 0 else None
        for cell in ws[ri]:
            cell.font      = d_font
            cell.alignment = d_align
            cell.border    = border
            if fill:
                cell.fill = fill

    # Column widths
    for ci, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
        hdr = col_cells[0].value or ""
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(hdr)) + 4, 12), 45)

    # Freeze header + auto-filter
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
