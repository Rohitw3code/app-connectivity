"""
RE Effectiveness PDF → Excel Pipeline (Integrated)
====================================================
1. Reads ALL PDFs from the 'effective' folder.
2. Extracts data using the SAME LLM configuration (gpt-4o-mini via llm_client).
3. Saves per-PDF extracted JSON in 'effectiveness_output/' — skips if already present.
4. Merges effectiveness data into the FINAL GNI output Excel by matching
   "GNA/ST II Application ID" → "Application ID".
5. Updates & adds columns as specified.
"""

import json
import os
import re
import sys
from pathlib import Path
from typing import Optional, Any

import pdfplumber
import pandas as pd
from pydantic import BaseModel, field_validator
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import MODEL, load_runtime_config
from llm_client import call_llm, extract_text_from_response
from data_extraction import extract_json_payload


# ─────────────────────────────────────────────────────────────
# 0. PATHS  (override with env vars or CLI args)
# ─────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent          # App-connectivity
EFFECTIVE_FOLDER = BASE_DIR / "CTUIL-Regenerators-Effective-Date-wise"
EFFECTIVENESS_OUTPUT_FOLDER = BASE_DIR / "start" / "effectiveness_output"

# ─────────────────────────────────────────────────────────────
# 1.  Pydantic schema for effectiveness records
# ─────────────────────────────────────────────────────────────
class RERecord(BaseModel):
    sl_no:                    Optional[str]   = None
    application_id:           Optional[str]   = None
    name_of_applicant:        Optional[str]   = None
    region:                   Optional[str]   = None
    type_of_project:          Optional[str]   = None
    installed_capacity_mw:    Optional[float] = None
    solar_mw:                 Optional[float] = None
    wind_mw:                  Optional[float] = None
    ess_mw:                   Optional[float] = None
    hydro_mw:                 Optional[float] = None
    connectivity_mw:          Optional[float] = None
    present_connectivity_mw:  Optional[float] = None
    substation:               Optional[str]   = None
    state:                    Optional[str]   = None
    expected_date:            Optional[str]   = None
    source_file:              Optional[str]   = None

    @field_validator(
        "installed_capacity_mw", "solar_mw", "wind_mw",
        "ess_mw", "hydro_mw", "connectivity_mw", "present_connectivity_mw",
        mode="before",
    )
    @classmethod
    def coerce_numeric(cls, v):
        if v in (None, "", "N/A", "-", "null", "—"):
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    @field_validator(
        "sl_no", "application_id", "name_of_applicant", "region",
        "type_of_project", "substation", "state", "expected_date",
        "source_file", mode="before",
    )
    @classmethod
    def coerce_str(cls, v):
        if v in (None, "", "N/A", "null"):
            return None
        s = str(v).replace("\n", " ").strip()
        return s if s else None


def safe_record(raw: dict) -> Optional[RERecord]:
    try:
        return RERecord(**raw)
    except Exception:
        clean = {k: raw.get(k) for k in RERecord.model_fields}
        try:
            return RERecord(**clean)
        except Exception:
            return None


# ─────────────────────────────────────────────────────────────
# 2.  LLM extraction prompt (effectiveness-specific)
# ─────────────────────────────────────────────────────────────
EFFECTIVENESS_SYSTEM_PROMPT = """You are a data extraction engine for Indian power sector regulatory PDFs \
(RE Effectiveness / Connectivity status reports).
Extract EVERY data row from the table text and return ONLY a valid JSON array.
Each element must be an object with these exact keys (null for missing):
  sl_no, application_id, name_of_applicant, region, type_of_project,
  installed_capacity_mw, solar_mw, wind_mw, ess_mw, hydro_mw, connectivity_mw,
  present_connectivity_mw, substation, state, expected_date

Rules:
- Skip header rows, footnotes, blank lines.
- Numeric fields must be float or null — never strings.
- expected_date: keep exactly as written e.g. "31-12-2025".
- application_id: the Application ID column (numeric ID).
- type_of_project: e.g. "Solar", "Wind", "Hybrid", "Solar + Wind", "Hydro", "ESS", etc.
- hydro_mw: capacity related to hydro/pump storage if present, else null.
- Output ONLY the JSON array — no prose, no markdown fences.
"""

EFFECTIVENESS_USER_TMPL = "Extract all rows:\n\n{text}"


def _build_effectiveness_prompt(page_text: str) -> dict[str, Any]:
    return {
        "messages": [
            {"role": "system", "content": EFFECTIVENESS_SYSTEM_PROMPT},
            {"role": "user",   "content": EFFECTIVENESS_USER_TMPL.format(text=page_text)},
        ],
        "temperature": 0,
        "max_tokens": 4000,
    }


# ─────────────────────────────────────────────────────────────
# 3.  LLM-based extraction for effectiveness PDFs
# ─────────────────────────────────────────────────────────────
def extract_effectiveness_with_llm(
    pdf_path: str,
    source_name: str,
    runtime,
) -> list[RERecord]:
    """Extract records from a single effectiveness PDF using the same LLM pipeline."""
    pages_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            if t.strip():
                pages_text.append(t)

    # Chunk pages to ~10 000 chars each
    chunks, cur, cur_len = [], [], 0
    for text in pages_text:
        if cur_len + len(text) > 10_000 and cur:
            chunks.append("\n\n".join(cur))
            cur, cur_len = [], 0
        cur.append(text)
        cur_len += len(text)
    if cur:
        chunks.append("\n\n".join(cur))

    records: list[RERecord] = []
    for i, chunk in enumerate(chunks, 1):
        print(f"  LLM chunk {i}/{len(chunks)} ...", end=" ", flush=True)
        for attempt in range(3):
            try:
                prompt_payload = _build_effectiveness_prompt(chunk)
                response_json = call_llm(
                    prompt_payload=prompt_payload,
                    vm=runtime.vm_mode,
                    api_key=runtime.api_key or None,
                    model=MODEL,
                    script_path=runtime.llm_script_path,
                )
                raw_text = extract_text_from_response(response_json)
                result = extract_json_payload(raw_text)

                if isinstance(result, list):
                    rows = result
                elif isinstance(result, dict):
                    rows = next(
                        (v for v in result.values() if isinstance(v, list)), []
                    )
                else:
                    rows = []

                batch = [
                    r for r in (
                        safe_record({**row, "source_file": source_name})
                        for row in rows
                    ) if r
                ]
                records.extend(batch)
                print(f"{len(batch)} rows")
                break
            except Exception as e:
                if attempt < 2:
                    import time
                    time.sleep(10)
                else:
                    print(f"FAILED ({e})")

    return records


# ─────────────────────────────────────────────────────────────
# 3b.  FALLBACK – pdfplumber table extraction
# ─────────────────────────────────────────────────────────────
_HEADER_MAP = {
    "si no":                              "sl_no",
    "sl. no.":                            "sl_no",
    "sl no":                              "sl_no",
    "application id":                     "application_id",
    "name of applicant":                  "name_of_applicant",
    "region":                             "region",
    "type of project":                    "type_of_project",
    "installed capacity (mw)":            "installed_capacity_mw",
    "installed capacity":                 "installed_capacity_mw",
    "solar":                              "solar_mw",
    "wind":                               "wind_mw",
    "ess":                                "ess_mw",
    "hydro":                              "hydro_mw",
    "connectivity (mw)":                  "connectivity_mw",
    "connectivity":                       "connectivity_mw",
    "present connectivity /deemed gna":   "present_connectivity_mw",
    "present connectivity":               "present_connectivity_mw",
    "substation":                         "substation",
    "state":                              "state",
    "expected date of connectivity/ gna to be made effective": "expected_date",
    "expected date of connectivity/gna to be made effective":  "expected_date",
    "expected date":                      "expected_date",
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower().strip())


def _map_headers(raw_headers: list) -> dict:
    mapping = {}
    for i, h in enumerate(raw_headers):
        key = _norm(h or "")
        if key in _HEADER_MAP:
            mapping[i] = _HEADER_MAP[key]
        else:
            for pattern, field in _HEADER_MAP.items():
                if key and pattern in key:
                    mapping[i] = field
                    break
    return mapping


def _is_header(row: list) -> bool:
    text = " ".join((c or "") for c in row).lower()
    return "application" in text or "sl. no" in text or "si no" in text


def extract_tables_fallback(pdf_path: str, source_name: str) -> list[RERecord]:
    records = []
    current_mapping: dict = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables({
                "vertical_strategy":   "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
            })
            if not tables:
                tbl = page.extract_table()
                tables = [tbl] if tbl else []

            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row or all(c is None for c in row):
                        continue
                    if _is_header(row):
                        current_mapping = _map_headers(row)
                        continue
                    if not current_mapping:
                        continue
                    raw = {}
                    for col_idx, field in current_mapping.items():
                        if col_idx < len(row):
                            raw[field] = row[col_idx]
                    if not raw.get("application_id") and not raw.get("name_of_applicant"):
                        continue
                    raw["source_file"] = source_name
                    rec = safe_record(raw)
                    if rec:
                        records.append(rec)

    return records


# ─────────────────────────────────────────────────────────────
# 4.  Deduplicate
# ─────────────────────────────────────────────────────────────
def deduplicate(records: list[RERecord]) -> list[RERecord]:
    seen, out = set(), []
    for r in records:
        key = (r.application_id or "").strip() + "||" + (r.name_of_applicant or "").strip()
        if key == "||":
            out.append(r)
            continue
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


# ─────────────────────────────────────────────────────────────
# 5.  Process ALL PDFs from the effective folder
# ─────────────────────────────────────────────────────────────
def _output_json_path(pdf_name: str) -> Path:
    """Return the path for the extracted JSON of a given PDF."""
    stem = Path(pdf_name).stem
    return EFFECTIVENESS_OUTPUT_FOLDER / f"{stem}.json"


def process_all_effectiveness_pdfs(
    effective_folder: Path | None = None,
    output_folder: Path | None = None,
) -> pd.DataFrame:
    """
    1. Scan the effective folder for all PDFs.
    2. For each PDF, check if output JSON already exists → skip if so.
    3. Extract using LLM (same config) and save JSON output.
    4. Load all JSON outputs and return a combined DataFrame.
    """
    global EFFECTIVE_FOLDER, EFFECTIVENESS_OUTPUT_FOLDER

    if effective_folder:
        EFFECTIVE_FOLDER = Path(effective_folder)
    if output_folder:
        EFFECTIVENESS_OUTPUT_FOLDER = Path(output_folder)

    EFFECTIVENESS_OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    if not EFFECTIVE_FOLDER.exists():
        print(f"[Effectiveness] Folder not found: {EFFECTIVE_FOLDER}")
        return pd.DataFrame()

    pdf_files = sorted(EFFECTIVE_FOLDER.glob("*.pdf"))
    if not pdf_files:
        print(f"[Effectiveness] No PDFs found in: {EFFECTIVE_FOLDER}")
        return pd.DataFrame()

    print(f"\n{'='*60}")
    print(f"  EFFECTIVENESS PIPELINE — {len(pdf_files)} PDFs found")
    print(f"{'='*60}")

    # Load runtime config (same as main pipeline)
    runtime = load_runtime_config()
    use_llm = bool(runtime.api_key)

    # --- Extract each PDF ---
    for pdf_path in pdf_files:
        source_name = pdf_path.name
        out_json = _output_json_path(source_name)

        if out_json.exists():
            print(f"  [SKIP] Already extracted: {source_name}")
            continue

        print(f"\n  [EXTRACT] {source_name}")
        if use_llm:
            records = extract_effectiveness_with_llm(
                str(pdf_path), source_name, runtime
            )
        else:
            records = extract_tables_fallback(str(pdf_path), source_name)

        records = deduplicate(records)
        print(f"  → {len(records)} unique records from {source_name}")

        # Serialize & save
        serialized = [r.model_dump() for r in records]
        out_json.parent.mkdir(parents=True, exist_ok=True)
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(serialized, f, indent=2, ensure_ascii=False)
        print(f"  → Saved: {out_json}")

    # --- Load all extracted JSONs → combined DataFrame ---
    all_records: list[RERecord] = []
    for json_file in sorted(EFFECTIVENESS_OUTPUT_FOLDER.glob("*.json")):
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            rec = safe_record(item)
            if rec:
                all_records.append(rec)

    all_records = deduplicate(all_records)
    print(f"\n  Total effectiveness records loaded: {len(all_records)}")

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame([r.model_dump() for r in all_records])
    return df


# ─────────────────────────────────────────────────────────────
# 6.  Merge effectiveness data into the final GNI Excel output
# ─────────────────────────────────────────────────────────────
def _is_valid(val) -> bool:
    """Check if a value is valid (not None, NaN, empty, or placeholder)."""
    if val is None:
        return False
    if isinstance(val, float) and pd.isna(val):
        return False
    s = str(val).strip().lower()
    return s not in ("", "none", "null", "na", "n/a", "-", "--", "nan")


def _classify_project_type(type_str: Optional[str]) -> dict:
    """
    From 'Type of Project', determine capacity breakdown categories.
    Returns dict with keys: solar, wind, ess, hydro, hybrid (bool/str).
    """
    if not type_str:
        return {}
    t = type_str.lower().strip()
    categories = []
    if "solar" in t:
        categories.append("solar")
    if "wind" in t:
        categories.append("wind")
    if "ess" in t or "energy storage" in t or "bess" in t:
        categories.append("ess")
    if "hydro" in t or "pump" in t or "psp" in t:
        categories.append("hydro")
    if len(categories) > 1 or "hybrid" in t:
        categories.append("hybrid")
    return {cat: True for cat in categories}


def merge_effectiveness_into_final(
    final_excel_path: str,
    effectiveness_df: pd.DataFrame,
    output_excel_path: str | None = None,
) -> str:
    """
    Read the final GNI output Excel, merge effectiveness data, and save.

    Matching: "GNA/ST II Application ID" (final) ↔ "application_id" (effectiveness)

    Updates (only if valid data present in effectiveness):
      - "Name of developers"               ← "name_of_applicant"
      - "Substation"                        ← "substation"  (mapped from 'substaion')
      - "State"                             ← "state"
      - "Application Quantum (MW)(ST II)"   ← "installed_capacity_mw"

    New columns added:
      - "Region"                            ← "region"
      - "type"                              ← "type_of_project"
      - "Installed capacity (MW) solar"     ← "solar_mw"
      - "Installed capacity (MW) wind"      ← "wind_mw"
      - "Installed capacity (MW) ess"       ← "ess_mw"
      - "Installed capacity (MW) hydro"     ← "hydro_mw"
      - "Installed capacity (MW) hybrid"    ← sum of breakdowns when hybrid
    """
    if output_excel_path is None:
        output_excel_path = final_excel_path

    if effectiveness_df.empty:
        print("[Merge] No effectiveness data to merge.")
        return final_excel_path

    # Read the final Excel
    final_df = pd.read_excel(final_excel_path, sheet_name=0)
    print(f"\n[Merge] Final Excel loaded: {len(final_df)} rows, columns: {list(final_df.columns)}")

    # --- Build lookup from effectiveness_df keyed on application_id ---
    eff_lookup: dict[str, dict] = {}
    for _, row in effectiveness_df.iterrows():
        app_id = str(row.get("application_id", "") or "").strip()
        if not app_id:
            continue
        # Keep latest (last encountered) entry per application_id
        eff_lookup[app_id] = row.to_dict()

    print(f"[Merge] Effectiveness lookup built: {len(eff_lookup)} unique application IDs")

    # --- Identify the GNA column in the final DataFrame ---
    gna_col = None
    for col in final_df.columns:
        if "GNA" in str(col) and "Application" in str(col) and "ID" in str(col):
            gna_col = col
            break
    if gna_col is None:
        # fallback
        gna_col = "GNA/ST II Application ID"

    # --- Determine column names in the final DF (handle variations) ---
    def _find_col(df, *candidates):
        for c in candidates:
            for col in df.columns:
                if c.lower() == col.lower():
                    return col
        return None

    col_name_dev = _find_col(final_df, "Name of the developers", "Name of developers")
    col_substation = _find_col(final_df, "substaion", "Substation")
    col_state = _find_col(final_df, "State")
    col_quantum = _find_col(final_df, "Application Quantum (MW)(ST II)")

    # --- Add new columns ---
    if "Region" not in final_df.columns:
        final_df["Region"] = None
    if "type" not in final_df.columns:
        final_df["type"] = None
    for sub_col in [
        "Installed capacity (MW) solar",
        "Installed capacity (MW) wind",
        "Installed capacity (MW) ess",
        "Installed capacity (MW) hydro",
        "Installed capacity (MW) hybrid",
    ]:
        if sub_col not in final_df.columns:
            final_df[sub_col] = None

    matched_count = 0
    for idx, row in final_df.iterrows():
        gna_ids_raw = str(row.get(gna_col, "") or "").strip()
        if not gna_ids_raw:
            continue

        # A row may have multiple IDs comma-separated
        gna_ids = [x.strip() for x in re.split(r"[,;\s]+", gna_ids_raw) if x.strip()]

        eff_data = None
        for gid in gna_ids:
            if gid in eff_lookup:
                eff_data = eff_lookup[gid]
                break

        if eff_data is None:
            continue

        matched_count += 1

        # --- Update columns (only if valid data in effectiveness) ---
        if col_name_dev and _is_valid(eff_data.get("name_of_applicant")):
            final_df.at[idx, col_name_dev] = eff_data["name_of_applicant"]

        if col_substation and _is_valid(eff_data.get("substation")):
            final_df.at[idx, col_substation] = eff_data["substation"]

        if col_state and _is_valid(eff_data.get("state")):
            final_df.at[idx, col_state] = eff_data["state"]

        if col_quantum and _is_valid(eff_data.get("installed_capacity_mw")):
            final_df.at[idx, col_quantum] = eff_data["installed_capacity_mw"]

        # --- New columns ---
        if _is_valid(eff_data.get("region")):
            final_df.at[idx, "Region"] = eff_data["region"]

        if _is_valid(eff_data.get("type_of_project")):
            final_df.at[idx, "type"] = eff_data["type_of_project"]

        # Installed capacity breakdown
        if _is_valid(eff_data.get("solar_mw")):
            final_df.at[idx, "Installed capacity (MW) solar"] = eff_data["solar_mw"]

        if _is_valid(eff_data.get("wind_mw")):
            final_df.at[idx, "Installed capacity (MW) wind"] = eff_data["wind_mw"]

        if _is_valid(eff_data.get("ess_mw")):
            final_df.at[idx, "Installed capacity (MW) ess"] = eff_data["ess_mw"]

        if _is_valid(eff_data.get("hydro_mw")):
            final_df.at[idx, "Installed capacity (MW) hydro"] = eff_data["hydro_mw"]

        # Hybrid = total of all breakdowns if type contains hybrid or multiple sources
        type_proj = eff_data.get("type_of_project", "") or ""
        cats = _classify_project_type(type_proj)
        if cats.get("hybrid"):
            hybrid_total = sum(
                float(eff_data.get(k) or 0)
                for k in ["solar_mw", "wind_mw", "ess_mw", "hydro_mw"]
                if _is_valid(eff_data.get(k))
            )
            if hybrid_total > 0:
                final_df.at[idx, "Installed capacity (MW) hybrid"] = hybrid_total

    print(f"[Merge] Matched & updated: {matched_count} rows out of {len(final_df)}")

    # --- Save ---
    Path(output_excel_path).parent.mkdir(parents=True, exist_ok=True)
    final_df.to_excel(output_excel_path, index=False, sheet_name="Extracted Data")

    # Apply formatting
    _format_output_excel(output_excel_path)
    print(f"[Merge] Saved merged output → {output_excel_path}")
    return output_excel_path


# ─────────────────────────────────────────────────────────────
# 7.  Excel formatting helper
# ─────────────────────────────────────────────────────────────
def _format_output_excel(excel_path: str):
    """Apply professional formatting to the output Excel."""
    wb = load_workbook(excel_path)
    ws = wb.active

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h_fill = PatternFill("solid", start_color="1F4E79")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d_font  = Font(name="Arial", size=9)
    d_align = Alignment(vertical="center", wrap_text=True)
    alt_fill = PatternFill("solid", start_color="EBF3FB")

    ws.row_dimensions[1].height = 40
    for cell in ws[1]:
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = h_align; cell.border = border

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in ws[row_idx]:
            cell.font = d_font; cell.alignment = d_align; cell.border = border
            if fill:
                cell.fill = fill

    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = col_cells[0].value or ""
        width = min(max(len(str(header)) + 4, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)


# ─────────────────────────────────────────────────────────────
# 8.  Standalone entry point
# ─────────────────────────────────────────────────────────────
def run_standalone(
    effective_folder: str | None = None,
    output_folder: str | None = None,
    final_excel_path: str | None = None,
    merged_output_path: str | None = None,
):
    """
    Run the full effectiveness pipeline:
    1. Extract all effectiveness PDFs (skip already extracted).
    2. Optionally merge into a final GNI Excel.
    """
    eff_folder = Path(effective_folder) if effective_folder else EFFECTIVE_FOLDER
    out_folder = Path(output_folder) if output_folder else EFFECTIVENESS_OUTPUT_FOLDER

    eff_df = process_all_effectiveness_pdfs(
        effective_folder=eff_folder,
        output_folder=out_folder,
    )

    if final_excel_path and Path(final_excel_path).exists():
        merged = merged_output_path or final_excel_path
        merge_effectiveness_into_final(final_excel_path, eff_df, merged)
    else:
        # Save effectiveness-only output
        if not eff_df.empty:
            eff_out = out_folder / "RE_effectiveness_combined.xlsx"
            eff_df.to_excel(str(eff_out), index=False)
            print(f"\n  Effectiveness-only Excel → {eff_out}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Effectiveness PDF extraction & merge pipeline")
    parser.add_argument(
        "--effective-folder",
        default=None,
        help=f"Folder containing effectiveness PDFs (default: {EFFECTIVE_FOLDER})",
    )
    parser.add_argument(
        "--output-folder",
        default=None,
        help=f"Folder to save extracted JSONs (default: {EFFECTIVENESS_OUTPUT_FOLDER})",
    )
    parser.add_argument(
        "--final-excel",
        default=None,
        help="Path to the final GNI Excel to merge effectiveness data into",
    )
    parser.add_argument(
        "--merged-output",
        default=None,
        help="Path for the merged output Excel (default: overwrite final-excel)",
    )
    args = parser.parse_args()

    run_standalone(
        effective_folder=args.effective_folder,
        output_folder=args.output_folder,
        final_excel_path=args.final_excel,
        merged_output_path=args.merged_output,
    )